from openpyxl import load_workbook,Workbook
import os,sys
file_path = sys.path[0]
sum = Workbook()
for fname in os.listdir(file_path):
    if fname.endswith('xlsx'):
        print(fname)
        src_worbook = load_workbook(fname)
        src_sheet = src_worbook.sheetnames[0]
        dest_sheet = sum.create_sheet(src_sheet)
        for i in range(1, src_sheet.max_row + 1):
            for j in range(1, src_sheet.max_column + 1):
                dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

sum.save('sum.xlsx')